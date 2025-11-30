"""
Daily statistics tracking and Excel report generation.
"""

import os
from datetime import datetime, timedelta
from collections import defaultdict
import threading
import time

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[Statistics] Warning: openpyxl not installed. Excel export disabled.")


class DailyStats:
    def __init__(self, report_dir='reports'):
        """
        Track hourly and daily statistics by gender, age group, and direction.
        
        Args:
            report_dir: Directory to save daily reports
        """
        self.report_dir = report_dir
        os.makedirs(report_dir, exist_ok=True)
        
        # Current date for tracking
        self.current_date = datetime.now().date()
        
        # Statistics storage
        self._init_stats()
        
        # Tracked person IDs to avoid double counting
        self.counted_ids = set()
        
        # Start background thread for auto-save
        self._stop_thread = False
        self._auto_save_thread = threading.Thread(target=self._auto_save_loop, daemon=True)
        self._auto_save_thread.start()
        
        print(f"[Statistics] Initialized. Reports will be saved to: {report_dir}")
    
    def _init_stats(self):
        """Initialize/reset statistics structures."""
        # Hourly counts: {hour: {"in": count, "out": count}}
        self.hourly_counts = defaultdict(lambda: {"in": 0, "out": 0})
        
        # Gender counts: {"Male": {"in": count, "out": count}, "Female": {...}}
        self.gender_counts = {
            "Male": {"in": 0, "out": 0},
            "Female": {"in": 0, "out": 0}
        }
        
        # Age group counts: {"<12": {"in": count, "out": count}, ...}
        self.age_counts = {
            "<12": {"in": 0, "out": 0},
            "13-25": {"in": 0, "out": 0},
            "26-45": {"in": 0, "out": 0},
            "46-60": {"in": 0, "out": 0},
            ">60": {"in": 0, "out": 0}
        }
        
        # Pending updates for people who crossed before analysis completed
        self.pending_updates = {}
        
        # Detailed hourly breakdown: {hour: {gender: {age_group: {"in": x, "out": y}}}}
        self.hourly_details = defaultdict(
            lambda: defaultdict(
                lambda: defaultdict(lambda: {"in": 0, "out": 0})
            )
        )
        
        # Total counts
        self.total_in = 0
        self.total_out = 0
        
        # Clear tracked IDs
        self.counted_ids = set()
    
    def _check_date_change(self):
        """Check if date changed and handle accordingly."""
        now = datetime.now()
        today = now.date()
        
        if today != self.current_date:
            # Date changed - save yesterday's report and reset
            yesterday = self.current_date
            print(f"[Statistics] Date changed from {yesterday} to {today}")
            
            # Save yesterday's report
            self.save_report(date_override=yesterday)
            
            # Reset for new day
            self.current_date = today
            self._init_stats()
            print(f"[Statistics] Stats reset for new day: {today}")
    
    def _auto_save_loop(self):
        """Background thread that saves report at 00:05 each day."""
        while not self._stop_thread:
            now = datetime.now()
            
            # Check if it's 00:05
            if now.hour == 0 and now.minute == 5:
                self._check_date_change()
                time.sleep(60)  # Wait a minute to avoid duplicate saves
            
            # Also check for date change periodically
            self._check_date_change()
            
            time.sleep(30)  # Check every 30 seconds
    
    def record_person(self, person_id, direction, gender=None, age_group=None):
        """
        Record a person crossing the line.
        
        Args:
            person_id: Unique ID of the person
            direction: "in" or "out"
            gender: "Male", "Female", or None
            age_group: "<12", "13-25", "26-45", "46-60", ">60", or None
        """
        # Create unique key for this crossing
        crossing_key = f"{person_id}_{direction}"
        
        if crossing_key in self.counted_ids:
            return  # Already counted this crossing
        
        self.counted_ids.add(crossing_key)
        
        # Get current hour
        hour = datetime.now().hour
        
        # Normalize inputs
        direction = direction.lower()
        
        # Update hourly counts (always)
        self.hourly_counts[hour][direction] += 1
        
        # Update totals (always)
        if direction == "in":
            self.total_in += 1
        else:
            self.total_out += 1
        
        # Only update gender/age if we have valid data (not None or Unknown)
        valid_gender = gender if gender in ["Male", "Female"] else None
        valid_age = age_group if age_group in ["<12", "13-25", "26-45", "46-60", ">60"] else None
        
        # Update gender counts only if valid
        if valid_gender:
            self.gender_counts[valid_gender][direction] += 1
        
        # Update age counts only if valid
        if valid_age:
            self.age_counts[valid_age][direction] += 1
        
        # Update detailed hourly breakdown only if we have both
        if valid_gender and valid_age:
            self.hourly_details[hour][valid_gender][valid_age][direction] += 1
        
        # Store pending update if missing data
        if not valid_gender or not valid_age:
            if not hasattr(self, 'pending_updates'):
                self.pending_updates = {}
            self.pending_updates[crossing_key] = {
                'hour': hour,
                'direction': direction,
                'gender': valid_gender,
                'age_group': valid_age
            }
    
    def update_pending(self, person_id, direction, gender=None, age_group=None):
        """
        Update pending record with gender/age data that was analyzed later.
        """
        crossing_key = f"{person_id}_{direction}"
        
        if not hasattr(self, 'pending_updates'):
            return
        
        if crossing_key not in self.pending_updates:
            return
        
        pending = self.pending_updates[crossing_key]
        hour = pending['hour']
        dir_lower = pending['direction']
        
        # Update gender if we now have it and didn't before
        if gender in ["Male", "Female"] and not pending['gender']:
            self.gender_counts[gender][dir_lower] += 1
            pending['gender'] = gender
        
        # Update age if we now have it and didn't before
        if age_group in ["<12", "13-25", "26-45", "46-60", ">60"] and not pending['age_group']:
            self.age_counts[age_group][dir_lower] += 1
            pending['age_group'] = age_group
        
        # Update detailed if we now have both
        if pending['gender'] and pending['age_group']:
            self.hourly_details[hour][pending['gender']][pending['age_group']][dir_lower] += 1
            del self.pending_updates[crossing_key]
    
    def get_current_stats(self):
        """Get current statistics summary."""
        return {
            "date": str(self.current_date),
            "total_in": self.total_in,
            "total_out": self.total_out,
            "inside": max(0, self.total_in - self.total_out),
            "gender": dict(self.gender_counts),
            "age": dict(self.age_counts),
            "hourly": dict(self.hourly_counts)
        }
    
    def save_report(self, date_override=None):
        """
        Save report as Excel file.
        
        Args:
            date_override: Optional date to use for filename (default: current_date)
        """
        if not OPENPYXL_AVAILABLE:
            print("[Statistics] Cannot save report: openpyxl not installed")
            return None
        
        report_date = date_override or self.current_date
        filename = f"report_{report_date}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        
        print(f"[Statistics] Saving report to: {filepath}")
        
        try:
            wb = Workbook()
            
            # === SUMMARY SHEET ===
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=12, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws_summary["A1"] = f"Daily Report - {report_date}"
            ws_summary["A1"].font = Font(bold=True, size=16)
            ws_summary.merge_cells("A1:D1")
            
            # Summary stats
            ws_summary["A3"] = "Total IN"
            ws_summary["B3"] = self.total_in
            ws_summary["A4"] = "Total OUT"
            ws_summary["B4"] = self.total_out
            ws_summary["A5"] = "Net Inside"
            ws_summary["B5"] = max(0, self.total_in - self.total_out)
            
            for row in range(3, 6):
                ws_summary[f"A{row}"].font = header_font
                ws_summary[f"B{row}"].alignment = center_align
            
            # Gender breakdown
            ws_summary["A7"] = "Gender Breakdown"
            ws_summary["A7"].font = Font(bold=True, size=14)
            ws_summary.merge_cells("A7:C7")
            
            ws_summary["A8"] = "Gender"
            ws_summary["B8"] = "IN"
            ws_summary["C8"] = "OUT"
            for col in ["A", "B", "C"]:
                ws_summary[f"{col}8"].font = header_font_white
                ws_summary[f"{col}8"].fill = header_fill
                ws_summary[f"{col}8"].alignment = center_align
            
            row = 9
            for gender, counts in self.gender_counts.items():
                # Skip Unknown category
                if gender == "Unknown":
                    continue
                if counts["in"] > 0 or counts["out"] > 0:
                    ws_summary[f"A{row}"] = gender
                    ws_summary[f"B{row}"] = counts["in"]
                    ws_summary[f"C{row}"] = counts["out"]
                    ws_summary[f"B{row}"].alignment = center_align
                    ws_summary[f"C{row}"].alignment = center_align
                    row += 1
            
            # Age breakdown
            age_start_row = row + 2
            ws_summary[f"A{age_start_row}"] = "Age Group Breakdown"
            ws_summary[f"A{age_start_row}"].font = Font(bold=True, size=14)
            ws_summary.merge_cells(f"A{age_start_row}:C{age_start_row}")
            
            age_start_row += 1
            ws_summary[f"A{age_start_row}"] = "Age Group"
            ws_summary[f"B{age_start_row}"] = "IN"
            ws_summary[f"C{age_start_row}"] = "OUT"
            for col in ["A", "B", "C"]:
                ws_summary[f"{col}{age_start_row}"].font = header_font_white
                ws_summary[f"{col}{age_start_row}"].fill = header_fill
                ws_summary[f"{col}{age_start_row}"].alignment = center_align
            
            row = age_start_row + 1
            for age_group in ["<12", "13-25", "26-45", "46-60", ">60"]:
                counts = self.age_counts[age_group]
                ws_summary[f"A{row}"] = age_group
                ws_summary[f"B{row}"] = counts["in"]
                ws_summary[f"C{row}"] = counts["out"]
                ws_summary[f"B{row}"].alignment = center_align
                ws_summary[f"C{row}"].alignment = center_align
                row += 1
            
            # Adjust column widths
            ws_summary.column_dimensions["A"].width = 18
            ws_summary.column_dimensions["B"].width = 12
            ws_summary.column_dimensions["C"].width = 12
            
            # === HOURLY BREAKDOWN SHEET ===
            ws_hourly = wb.create_sheet("Hourly Breakdown")
            
            # Headers
            ws_hourly["A1"] = "Hour"
            ws_hourly["B1"] = "IN"
            ws_hourly["C1"] = "OUT"
            ws_hourly["D1"] = "Net"
            for col in ["A", "B", "C", "D"]:
                ws_hourly[f"{col}1"].font = header_font_white
                ws_hourly[f"{col}1"].fill = header_fill
                ws_hourly[f"{col}1"].alignment = center_align
            
            # Data
            for hour in range(24):
                row = hour + 2
                counts = self.hourly_counts[hour]
                ws_hourly[f"A{row}"] = f"{hour:02d}:00"
                ws_hourly[f"B{row}"] = counts["in"]
                ws_hourly[f"C{row}"] = counts["out"]
                ws_hourly[f"D{row}"] = counts["in"] - counts["out"]
                for col in ["A", "B", "C", "D"]:
                    ws_hourly[f"{col}{row}"].alignment = center_align
            
            ws_hourly.column_dimensions["A"].width = 10
            ws_hourly.column_dimensions["B"].width = 10
            ws_hourly.column_dimensions["C"].width = 10
            ws_hourly.column_dimensions["D"].width = 10
            
            # === PIE CHART DATA SHEET ===
            ws_pie = wb.create_sheet("Charts Data")
            
            # Gender pie data (exclude Unknown)
            ws_pie["A1"] = "Gender Distribution (IN)"
            ws_pie["A1"].font = Font(bold=True, size=12)
            ws_pie["A2"] = "Gender"
            ws_pie["B2"] = "Count"
            ws_pie["A2"].font = header_font
            ws_pie["B2"].font = header_font
            
            row = 3
            for gender in ["Male", "Female"]:
                count = self.gender_counts.get(gender, {}).get("in", 0)
                if count > 0:
                    ws_pie[f"A{row}"] = gender
                    ws_pie[f"B{row}"] = count
                    row += 1
            
            # Age pie data
            age_pie_start = row + 2
            ws_pie[f"A{age_pie_start}"] = "Age Distribution (IN)"
            ws_pie[f"A{age_pie_start}"].font = Font(bold=True, size=12)
            ws_pie[f"A{age_pie_start + 1}"] = "Age Group"
            ws_pie[f"B{age_pie_start + 1}"] = "Count"
            ws_pie[f"A{age_pie_start + 1}"].font = header_font
            ws_pie[f"B{age_pie_start + 1}"].font = header_font
            
            row = age_pie_start + 2
            for age_group in ["<12", "13-25", "26-45", "46-60", ">60"]:
                count = self.age_counts.get(age_group, {}).get("in", 0)
                if count > 0:
                    ws_pie[f"A{row}"] = age_group
                    ws_pie[f"B{row}"] = count
                    row += 1
            
            ws_pie.column_dimensions["A"].width = 15
            ws_pie.column_dimensions["B"].width = 10
            
            # Save workbook
            wb.save(filepath)
            print(f"[Statistics] Report saved: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"[Statistics] Error saving report: {e}")
            return None
    
    def stop(self):
        """Stop the auto-save thread."""
        self._stop_thread = True
        if self._auto_save_thread.is_alive():
            self._auto_save_thread.join(timeout=2)

